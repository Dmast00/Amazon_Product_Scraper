from typing import Text
import openpyxl
from openpyxl.cell import cell
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet import worksheet
from openpyxl.styles import Alignment, alignment
from selenium import webdriver
from bs4 import BeautifulSoup
from openpyxl import workbook,load_workbook
from os import cpu_count, startfile,close
from selenium.common import exceptions
from selenium.webdriver.common import keys
from selenium.webdriver.firefox import options 
from selenium.webdriver.support import wait
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
from selenium.common.exceptions import InvalidSessionIdException, NoSuchElementException, StaleElementReferenceException, TimeoutException
from selenium.webdriver.common.keys import Keys
import pandas as pd 
import re
import time
import os
from time  import time,sleep
Menu_Options ={
    1:'Search From Excel',
    2:'Search Only one Product',
    3:'Exit'
}
url = "https://amazon.com"
def Print_Menu():
    for key in Menu_Options:
        print(key,'--',Menu_Options[key])

def Count_items(driver):
    items =driver.find_elements(By.CSS_SELECTOR,'.s-main-slot div.s-result-item.s-asin.sg-col-0-of-12.sg-col-16-of-20.sg-col')
    # Search for the Asin code and opens the product details page
    products_len =len(items) 
    return products_len

def Search_From_Excel():
    
    path_temp = str(input('file name:'))
    path = path_temp+'.xlsx'
    workbook = Workbook()
    wb = load_workbook(path)
    sheet = wb.active
    driver = webdriver.Firefox()
    count = 1
    
    for cell in sheet['A']:
        prod_name = cell.value
        Excel_new_sheet = prod_name+'.xlsx'
        wb = openpyxl.Workbook(Excel_new_sheet)
        wb.save(Excel_new_sheet)
        driver.get(url)
        
        search_bar = driver.find_element_by_id('twotabsearchtextbox')
        search_bar.send_keys(prod_name)

        search_button = driver.find_element_by_xpath('//*[@id="nav-search-submit-button"]')
        search_button.click()

        prods_len = Count_items(driver)

        i = 1
        count = 1
        exc_cel = 1 
        while count <= prods_len:
            
            sleep(2)
            try:
                search_asin = driver.find_element_by_xpath('/html/body/div[1]/div[2]/div[1]/div[1]/div/span[3]/div[2]/div[{}]'.format(i))                                      
                asin = search_asin.get_attribute('data-asin')

                if asin == '':
                    i +=1
                    continue                                     
            except:

                print('ASIN not found')                                                                           
            
            sleep(2)
            try:
                product = driver.find_element_by_xpath('/html/body/div[1]/div[2]/div[1]/div[1]/div/span[3]/div[2]/div[{}]/div/span/div/div/div[2]/div[2]/div/div/div[1]/h2/a/span'.format(i)).text
                                        
            except:
                product = driver.find_element_by_css_selector('.widgetId\=search-results_{} > div:nth-child(1) > div:nth-child(1) > div:nth-child(2) > div:nth-child(2) > div:nth-child(1) > div:nth-child(1) > div:nth-child(1) > h2:nth-child(1) > a:nth-child(1) > span:nth-child(1)'.format(i)).text
            try:
                if product is None:
                    product = driver.find_element_by_css_selector('.widgetId\=search-results_{} > div:nth-child(1) > div:nth-child(1) > div:nth-child(1) > div:nth-child(1) > div:nth-child(2) > div:nth-child(2) > div:nth-child(1) > div:nth-child(1) > div:nth-child(1) > h2:nth-child(2) > a:nth-child(1) > span:nth-child(1)'.format(i)).text
                elif product is None:
                    continue
                try:
                    sheet_name = product[0:25]
                    sleep(1)
                    workbook = load_workbook(filename=Excel_new_sheet)
                    workbook.create_sheet(title=sheet_name)
                    workbook['{}'.format(sheet_name)]['A{}'.format(exc_cel)] = product
                    workbook.save(Excel_new_sheet)
                    count +=1
                    i+=1
                    exc_cel+=1

                   

                except:
                    pass
            except:
                driver.back()






















def Count_review_items(driver):
    # reviews = driver.find_elements(By.CSS_SELECTOR,'.a-section.review.aok-relative')
    # reviews = driver.find_elements(By.CSS_SELECTOR,'div.a-section.celwidget')
    try:
        reviews = driver.find_elements(By.CSS_SELECTOR,'div.a-section.review.aok-relative')
        reviews_id = []
        for rev in reviews:
            reviews_id.append(rev.get_attribute('id'))
        reviews_len = len(reviews)
        print(reviews_id)
        print('Counting Function is Ok')
        
        return reviews_id
    except:
        return
    


def Select_Product_Name():
    cmd = 'cls'
    os.system(cmd)
    product_name = ''
    product_name = str(input('Enter the name of the prodcut: '))
    excel_name = product_name+'.xlsx'
    # driver = webdriver.Firefox()
    xl = openpyxl.Workbook(excel_name)
    xl.save(excel_name)
    # driver.get(url)
    product= ''
    driver = webdriver.Firefox()
    driver.get(url)

    search_bar = driver.find_element_by_id('twotabsearchtextbox')
    search_bar.send_keys(product_name)

    search_button = driver.find_element_by_xpath('//*[@id="nav-search-submit-button"]')
    search_button.click()
    count = 1
    Search_Only_One_Product(driver,product_name,count)
    return product_name

def Search_reviews(driver,wbFileName,count,Excel_sheet_name):
    try:
        workbook = Workbook()
        workbook = load_workbook(filename=wbFileName)
        sheet = workbook.active
        id_reviews = Count_review_items(driver)
        
        for id in id_reviews:
            try:
                rating_temp = driver.find_element_by_css_selector('#customer_review-{} > div:nth-child(2) > a:nth-child(1)'.format(id))
            except:
                rating_temp =driver.find_element_by_css_selector('#customer_review_foreign-{}} > div:nth-child(2) > i:nth-child(1)'.format(id))
                                                               
            sleep(2)
            rating = rating_temp.get_attribute('title')
            comment = driver.find_element_by_css_selector('#customer_review-{} > div:nth-child(5) > span:nth-child(1) > span:nth-child(1)'.format(id)).text
                                                                                        
            if comment is None:
                comment = driver.find_element_by_css_selector('#customer_review-{} > div:nth-child(5) > span:nth-child(1) > span:nth-child(4)'.format(id)).text
            sleep(2)
            print(comment)
            workbook['{}'.format(Excel_sheet_name)]['D{}'.format(count)] = rating
            sleep(1)
            workbook.save(wbFileName)
            # workbook['{}'.format(Excel_sheet_name)]['C'].alignment = Alignment(wrap_text=True)
            workbook['{}'.format(Excel_sheet_name)]['C{}'.format(count)] = comment
            sleep(1)
            workbook.save(wbFileName)
            count+=1
            sleep(3)
        try:
            next_page_review = driver.find_element_by_css_selector('.a-last > a:nth-child(1)')
                                                                   
            sleep(2)                                                  
            next_page_review.click()
            sleep(3)
        except:
            pass
        if next_page_review is None:
            print('No more pages available')
            return
        Search_reviews(driver,wbFileName,count,Excel_sheet_name)
    except:
        pass
    return
    
def Search_Only_One_Product(driver,product_name,count):
    sleep(2)
    Products_len = Count_items(driver)
    wbFileName = product_name+''+'.xlsx'
    workbook = Workbook()
    workbook = load_workbook(filename=wbFileName)
    workbook.save(wbFileName)
    print(wbFileName)
    sheet = workbook.active
    count = 1
    i = 1
    print(f'Count: {count}')
    while i <= Products_len:
        count = 1 
        sleep(2)

        try:
            search_asin = driver.find_element_by_xpath('/html/body/div[1]/div[2]/div[1]/div[1]/div/span[3]/div[2]/div[{}]'.format(i))
                                                        
            asin = search_asin.get_attribute('data-asin')
            if asin == '':
                i +=1
                continue                                     
        except:                                         
            print('ASIN not found')                                                                           
        
        sleep(2)
        try:
            product = driver.find_element_by_xpath('/html/body/div[1]/div[2]/div[1]/div[1]/div/span[3]/div[2]/div[{}]/div/span/div/div/div[2]/div[2]/div/div/div[1]/h2/a/span'.format(i)).text
                                       
        except:
            product = driver.find_element_by_css_selector('.widgetId\=search-results_{} > div:nth-child(1) > div:nth-child(1) > div:nth-child(2) > div:nth-child(2) > div:nth-child(1) > div:nth-child(1) > div:nth-child(1) > h2:nth-child(1) > a:nth-child(1) > span:nth-child(1)'.format(i)).text
                                                                                                    
        try:
            if product is None:
                product = driver.find_element_by_css_selector('.widgetId\=search-results_{} > div:nth-child(1) > div:nth-child(1) > div:nth-child(1) > div:nth-child(1) > div:nth-child(2) > div:nth-child(2) > div:nth-child(1) > div:nth-child(1) > div:nth-child(1) > h2:nth-child(2) > a:nth-child(1) > span:nth-child(1)'.format(i)).text
                                                                    
            elif product is None:
                continue
            try:
                                                    
                
                sleep(1)
                Excel_sheet_name = product[0:20]
                print(Excel_sheet_name)

                workbook = load_workbook(filename=wbFileName)
                workbook.create_sheet(title=Excel_sheet_name)
                workbook['{}'.format(Excel_sheet_name)]['A{}'.format(count)] = product
                workbook.save(wbFileName)

                if asin is not None:
                    new_window = url+'/dp/'+asin
                    driver.get(new_window)
                    try:
                        workbook['{}'.format(Excel_sheet_name)]['B{}'.format(count)] = new_window
                        workbook.save(wbFileName)
                        reviews = driver.find_element_by_css_selector('#reviews-medley-footer > div:nth-child(2) > a:nth-child(2)')                                           
                        if reviews is None:
                            print('no comments section')
                        else:                                         
                            rev_url = reviews.get_attribute('href')
                            driver.execute_script("window.open('{}');".format(rev_url))
                            driver.switch_to.window(driver.window_handles[1])
                            sleep(4)
                            Search_reviews(driver,wbFileName,count,Excel_sheet_name)
                            driver.close()
                            driver.switch_to.window(driver.window_handles[0])
                            driver.back()
                    except:
                        driver.back()
                else:
                    driver.back()
            except:
                pass
        except:
            print(i)
            driver.back()
        
        
        i +=1
        sleep(2)
        count +=1
    print(count)
    try:
        next_page = driver.find_element_by_css_selector('li.a-last > a:nth-child(1)')                                                        
        next_page.click()
        Search_Only_One_Product(driver,product_name,count)
        
    except:
        pass
    try:
        next_page = driver.find_element_by_css_selector('a.s-pagination-item:nth-child(7)')                                                      
        next_page.click()
        Search_Only_One_Product(driver,product_name,count)
    except:
        return


def Closing_Script():
    exit()


if __name__ == '__main__':
    while (True):
        Print_Menu()
        option =''
        try:
            option = int(input('Enter an Option: '))
        except:
            print('Wrong input please enter a number!!')
        if option == 1:
                Search_From_Excel()
        elif option == 2:
                Select_Product_Name()
        elif option == 3:
                Closing_Script()
        else:
            print('Please enter a number between 1 and 3!!')



